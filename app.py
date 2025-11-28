import streamlit as st
import pandas as pd
import numpy as np
import re
import json
import math
from io import BytesIO
from openpyxl import load_workbook

st.set_page_config(page_title="Tender Pricing App", layout="wide")

# ---------- Defaults meta (for colouring old vs new groups) ----------
DEFAULT_GROUP_NAMES = {
    "CORFLUTE_3MM",
    "SCREENBOARD_2MM",
    "POSTER_BOARD",
    "WINDOW_SUPERCLING",
    "BANNER_SYNTHETIC",
    "FERROUS",
    "VINYL_MPI1105",
    "VINYL_3M7725",
    "VINYL_DIGITAL",
    "VINYL_PLOTTER",
    "ACM",
    "ACRYLIC",
    "BRAILLE",
    "STICKERS",
    "OTHER",
}

DEFAULT_DS_MATERIALS = {
    "Synthetic Banner Double Sided",
    "Banner Synthetic DS",
    "Banner Synthetic Double Sided",
}

DEFAULT_DS_LOADING = 0.25  # 25%

# For saving + restoring mappings etc.
STATE_VERSION = 5


# ---------- Utility helpers ----------


def safe_json_loads(s, default=None):
    if s is None:
        return default
    try:
        return json.loads(s)
    except Exception:
        return default


def infer_letter_from_header(header: str) -> str | None:
    """
    Try to extract Excel column letter from a header like 'AC - Size' or 'AC Size'.
    If header is literally a single or double letters, returns that.
    """
    if not isinstance(header, str):
        return None

    header = header.strip().upper()

    # If header itself is A, B, AA, AB etc
    if re.fullmatch(r"[A-Z]{1,3}", header):
        return header

    # Look for pattern like 'AC -' or 'AC '
    m = re.match(r"([A-Z]{1,3})\\b", header)
    if m:
        return m.group(1)

    return None


def col_letter_to_index(letter: str) -> int:
    """Convert Excel-style column letter (A,B,...,AA,AB,...) to zero-based index."""
    if not isinstance(letter, str) or not letter:
        raise ValueError("Invalid column letter")

    letter = letter.strip().upper()
    result = 0
    for ch in letter:
        if not ("A" <= ch <= "Z"):
            raise ValueError(f"Invalid column letter: {letter}")
        result = result * 26 + (ord(ch) - ord("A") + 1)
    return result - 1


def col_index_to_letter(idx: int) -> str:
    """Convert zero-based index to Excel-style column letter."""
    if idx < 0:
        raise ValueError("Negative index not allowed")
    idx += 1
    letters = []
    while idx:
        idx, rem = divmod(idx - 1, 26)
        letters.append(chr(rem + ord("A")))
    return "".join(reversed(letters))


def parse_dimension_to_sqm(dim_str: str) -> float:
    """
    Parse strings like:
    - '841mm x 1189mm', '594 x 841mm', '1.2m x 2m'
    - '260mm Ø', '50mm Diameter', '290mm'  (treated as diameter)
    - '2 x 2547mm x 755mm 2 x 967mm x 755mm'
    - 'Various - 7 Decals 1. 2 x 355mm x 355mm 2. 2 x 120mm x 170mm ... 7. 4 x 30mm x 30mm'
    - '25,000mm (w) x 75mm (h)   10 strips x 2,500mm (w)'

    Returns SQM per row "set":
    - for normal rectangles: w × h
    - for circles: π × (d/2)²
    - for multi-size strings: sum over all (qty × w × h)
    """
    if pd.isna(dim_str):
        return np.nan

    s = str(dim_str).strip().lower()
    if not s:
        return np.nan

    # Normalise separators
    s = s.replace("×", "x")

    # Remove thousand separators inside numbers, e.g. 25,000mm -> 25000mm
    s = re.sub(r"(?<=\\d),(?=\\d)", "", s)

    def to_m(v, u):
        """Convert numeric value + unit to metres."""
        if u == "cm":
            return v / 100.0
        if u == "m":
            return v
        # default (mm)
        return v / 1000.0

    # 1) Explicit diameter formats: Ø, "diameter", "dia"
    if any(tok in s for tok in ["diameter", "ø", "⌀", " dia", "dia "]):
        m = re.search(r"(\\d+(\\.\\d+)?)\\s*(mm|cm|m)?", s)
        if m:
            v = float(m.group(1))
            unit = m.group(3) or "mm"
            d_m = to_m(v, unit)
            return math.pi * (d_m / 2.0) ** 2

    # 2) Multi-size / "various" patterns with explicit qty:
    #    e.g. '2 x 355mm x 355mm', '4 x 70mm x 60mm'
    pattern_q = re.compile(
        r"(?P<qty>\\d+)\\s*x\\s*"
        r"(?P<w>\\d+(\\.\\d+)?)\\s*(?P<uw>mm|cm|m)?\\s*x\\s*"
        r"(?P<h>\\d+(\\.\\d+)?)\\s*(?P<uh>mm|cm|m)?"
    )

    total_area = 0.0
    any_match = False
    for m in pattern_q.finditer(s):
        any_match = True
        qty = int(m.group("qty"))
        w = float(m.group("w"))
        h = float(m.group("h"))
        uw = m.group("uw") or "mm"
        uh = m.group("uh") or "mm"
        w_m = to_m(w, uw)
        h_m = to_m(h, uh)
        total_area += qty * w_m * h_m

    if any_match:
        return total_area

    # 3) Standard rectangle: first "w x h" pair
    pattern_wh = re.compile(
        r"(?P<w>\\d+(\\.\\d+)?)\\s*(?P<uw>mm|cm|m)?\\s*x\\s*"
        r"(?P<h>\\d+(\\.\\d+)?)\\s*(?P<uh>mm|cm|m)?"
    )
    m = pattern_wh.search(s)
    if m:
        w = float(m.group("w"))
        h = float(m.group("h"))
        uw = m.group("uw") or "mm"
        uh = m.group("uh") or "mm"
        w_m = to_m(w, uw)
        h_m = to_m(h, uh)
        return w_m * h_m

    # 4) Single numeric value: treat as diameter
    nums = re.findall(r"(\\d+(\\.\\d+)?)", s)
    if len(nums) == 1:
        v = float(nums[0][0])
        um = re.search(r"(mm|cm|m)", s)
        unit = um.group(1) if um else "mm"
        d_m = to_m(v, unit)
        return math.pi * (d_m / 2.0) ** 2

    # 5) Fallback: first two numbers as width & height rectangle
    matches = re.findall(r"(\\d+(\\.\\d+)?)\\s*(mm|cm|m)?", s)
    if len(matches) < 2:
        return np.nan

    (v1, _, u1) = matches[0]
    (v2, _, u2) = matches[1]
    v1 = float(v1)
    v2 = float(v2)
    u1 = u1 or "mm"
    u2 = u2 or "mm"

    w = to_m(v1, u1)
    h = to_m(v2, u2)
    return w * h


def detect_side(text, ds_synonyms, ss_synonyms, default="SS"):
    """Return 'DS' or 'SS' based on synonyms found in text."""
    if pd.isna(text):
        return default

    s = str(text).strip()
    s_lower = s.lower()

    # If it literally already contains DS or SS tokens, prefer that
    if "ds" in s_lower and "ss" not in s_lower:
        return "DS"
    if "ss" in s_lower and "ds" not in s_lower:
        return "SS"

    for word in ds_synonyms:
        if word.lower() in s_lower:
            return "DS"
    for word in ss_synonyms:
        if word.lower() in s_lower:
            return "SS"

    return default


def normalise_material_name(name: str | None) -> str | None:
    """Simple normalisation to help matching materials."""
    if name is None or (isinstance(name, float) and pd.isna(name)):
        return None
    s = str(name).strip()
    if not s:
        return None
    return " ".join(s.split())


def colour_for_group(group_name: str, existing_groups: set[str]) -> str:
    """
    Return background colour for a group name:
    - greyish for default meta groups
    - light green for new groups
    """
    if group_name in DEFAULT_GROUP_NAMES:
        return "#f0f0f0"
    if group_name in existing_groups:
        return "#ffffff"
    return "#e5ffe5"


def convert_letter_config_to_headers(df: pd.DataFrame, cfg: dict) -> dict:
    """
    Convert config referencing columns by letter (e.g. 'O', 'P') into
    config referencing the actual df columns by header name.
    """
    letter_to_header = {col_index_to_letter(i): c for i, c in enumerate(df.columns)}

    out = cfg.copy()
    for key in [
        "size_source_letter",
        "qty_annum_source_letter",
        "side_source_letter",
        "group_source_letter",
        "material_source_letter",
        "qty_per_period_source_letter",
    ]:
        letter = cfg.get(key)
        if letter:
            header = letter_to_header.get(letter)
            out[key.replace("_letter", "")] = header
        else:
            out[key.replace("_letter", "")] = None
    return out


def convert_headers_to_letter_config(df: pd.DataFrame, cfg: dict) -> dict:
    """
    Inverse of convert_letter_config_to_headers.
    """
    header_to_letter = {c: col_index_to_letter(i) for i, c in enumerate(df.columns)}
    out = cfg.copy()

    mapping = {
        "size_source": "size_source_letter",
        "qty_annum_source": "qty_annum_source_letter",
        "side_source": "side_source_letter",
        "group_source": "group_source_letter",
        "material_source": "material_source_letter",
        "qty_per_period_source": "qty_per_period_source_letter",
    }

    for header_key, letter_key in mapping.items():
        header = cfg.get(header_key)
        if header:
            out[letter_key] = header_to_letter.get(header)
        else:
            out[letter_key] = None

    return out


# ---------- Core pricing logic ----------


def build_pricing_rows(
    df: pd.DataFrame,
    size_col: str | None,
    mat_col: str | None,
    qty_annum_col: str | None,
    side_mode: str,
    side_fixed_value: str,
    side_embedded_pattern: str,
    side_source_letter: str | None,
    qty_annum_default: float | None,
    default_side: str,
    group_mode: str,
    group_fixed_name: str | None,
    group_source_letter: str | None,
    ds_materials: set[str],
    ds_loading: float,
    material_groups: dict[str, list[str]],
    material_rates: dict[str, float],
    qty_per_period_col: str | None,
):
    """
    Walk DF rows, parse size, choose group, side, sqm and price.
    Returns a list of (index, result_dict).
    """

    results = []

    # Build synonyms for DS/SS detection from DS materials and some defaults
    ds_synonyms = {"double", "double sided", "ds", "two sided", "2 sided"}
    ss_synonyms = {"single", "single sided", "ss", "one sided", "1 sided"}

    # Add group names as hints, e.g. "Banner Synthetic DS"
    for m in ds_materials:
        ds_synonyms.add(m.lower())

    letter_to_header = {col_index_to_letter(i): c for i, c in enumerate(df.columns)}

    group_src_col = (
        letter_to_header.get(group_source_letter)
        if group_mode == "From another column" and group_source_letter
        else None
    )
    side_src_col = (
        letter_to_header.get(side_source_letter)
        if side_mode == "Embedded in another column" and side_source_letter
        else None
    )

    for idx, row in df.iterrows():
        size_val = row[size_col] if size_col else None
        material_val = row[mat_col] if mat_col else None

        qty_annum = (
            pd.to_numeric(row[qty_annum_col], errors="coerce") if qty_annum_col else np.nan
        )

        if qty_per_period_col:
            # If qty per period is given, that is the main qty; if not numeric, fall back
            main_qty = pd.to_numeric(row[qty_per_period_col], errors="coerce")
            if pd.isna(main_qty):
                main_qty = qty_annum
        else:
            main_qty = qty_annum

        if pd.isna(main_qty):
            if qty_annum_default is not None:
                main_qty = qty_annum_default
            else:
                main_qty = np.nan

        # Determine side
        if side_mode == "Fixed":
            side = side_fixed_value
        elif side_mode == "Embedded in another column" and side_src_col:
            side_text = row[side_src_col]
            side = detect_side(side_text, ds_synonyms, ss_synonyms, default=default_side)
        else:
            side = default_side

        # Determine group
        if group_mode == "Fixed":
            group_name = group_fixed_name or "UNGROUPED"
        elif group_mode == "From another column" and group_src_col:
            group_name = normalise_material_name(row[group_src_col]) or "UNGROUPED"
        else:
            group_name = "UNGROUPED"

        # Determine matched material group label
        material_norm = normalise_material_name(material_val)
        matched_group = None

        for group_label, materials in material_groups.items():
            for pattern in materials:
                if material_norm and pattern.lower() in material_norm.lower():
                    matched_group = group_label
                    break
            if matched_group:
                break

        if matched_group:
            group_name = matched_group

        # Parse size -> sqm
        if size_val is not None:
            sqm_each = parse_dimension_to_sqm(size_val)
        else:
            sqm_each = np.nan

        # Calculate total sqm
        if pd.isna(main_qty) or pd.isna(sqm_each):
            total_sqm = np.nan
        else:
            total_sqm = float(main_qty) * float(sqm_each)

        # Find rate from group_name
        rate = material_rates.get(group_name, np.nan)

        # DS loading if relevant
        is_ds_material = group_name in ds_materials
        if side == "DS" and is_ds_material and not pd.isna(rate):
            eff_rate = rate * (1 + ds_loading)
        else:
            eff_rate = rate

        if pd.isna(eff_rate) or pd.isna(total_sqm):
            total_price = np.nan
        else:
            total_price = eff_rate * total_sqm

        results.append(
            (
                idx,
                {
                    "Size": size_val,
                    "Material": material_val,
                    "Group": group_name,
                    "Side": side,
                    "Qty": main_qty,
                    "SQM_each": sqm_each,
                    "Total_SQM": total_sqm,
                    "Rate_per_SQM": eff_rate,
                    "Total_Price": total_price,
                },
            )
        )

    return results


def write_results_to_workbook(
    wb_bytes: bytes,
    df_original: pd.DataFrame,
    results: list,
    sqm_col_letter: str,
    price_col_letter: str,
    sqm_row_offset: int,
    price_row_offset: int,
    start_data_row: int,
):
    """Write SQM and Price data back into the workbook."""
    wb = load_workbook(BytesIO(wb_bytes), data_only=False)
    ws = wb.active

    sqm_col_idx = col_letter_to_index(sqm_col_letter) + 1
    price_col_idx = col_letter_to_index(price_col_letter) + 1

    # We assume df_original index corresponds to worksheet row offset by start_data_row
    for idx, res in results:
        excel_row = start_data_row + idx  # currently unused but kept for future layout changes

        sqm_cell = ws.cell(row=sqm_row_offset, column=sqm_col_idx + idx)
        price_cell = ws.cell(row=price_row_offset, column=price_col_idx + idx)

        sqm_val = res["Total_SQM"]
        price_val = res["Total_Price"]

        if isinstance(sqm_val, (float, int)) and not math.isnan(sqm_val):
            sqm_cell.value = float(sqm_val)
        else:
            sqm_cell.value = None

        if isinstance(price_val, (float, int)) and not math.isnan(price_val):
            price_cell.value = float(price_val)
        else:
            price_cell.value = None

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


# ---------- Streamlit UI ----------


def main():
    st.title("Tender Pricing App")

    st.markdown(
        """
        Upload your Excel tender sheet, map the important columns, define material groups and rates,
        then generate an updated workbook with SQM and pricing written back in.
        """
    )

    uploaded = st.file_uploader("Upload Excel template", type=["xlsx"])

    if not uploaded:
        st.info("Upload an Excel file to begin.")
        return

    file_bytes = uploaded.read()

    # Read workbook as dataframe (first sheet)
    try:
        df = pd.read_excel(BytesIO(file_bytes), sheet_name=0, header=0, dtype=str)
    except Exception as e:
        st.error(f"Could not read Excel file: {e}")
        return

    st.subheader("Preview of input data")
    st.dataframe(df.head(50))

    columns = list(df.columns)

    st.sidebar.header("1. Map size / material / qty columns")

    size_col = st.sidebar.selectbox("Size column", ["(none)"] + columns, index=1 if len(columns) > 1 else 0)
    mat_col = st.sidebar.selectbox("Material column", ["(none)"] + columns, index=2 if len(columns) > 2 else 0)

    qty_annum_col = st.sidebar.selectbox(
        "Qty per annum column (or main qty if only one)",
        ["(none)"] + columns,
        index=3 if len(columns) > 3 else 0,
    )
    qty_annum_default = st.sidebar.number_input(
        "Default qty if blank", min_value=0.0, value=0.0, step=1.0, help="If 0, blanks stay blank."
    )
    if qty_annum_default == 0:
        qty_annum_default_val = None
    else:
        qty_annum_default_val = qty_annum_default

    qty_per_period_col = st.sidebar.selectbox(
        "Optional: Qty per period column (takes precedence as main Qty)",
        ["(none)"] + columns,
        index=0,
    )
    if qty_per_period_col == "(none)":
        qty_per_period_col = None

    # ---------- Side settings ----------
    st.sidebar.header("2. Side (SS / DS)")

    side_mode = st.sidebar.radio("Side mode", ["Default only", "Fixed", "Embedded in another column"])

    default_side = st.sidebar.selectbox("Default side", ["SS", "DS"], index=0)

    side_fixed_value = "SS"
    side_source_letter = None
    side_embedded_pattern = ""

    if side_mode == "Fixed":
        side_fixed_value = st.sidebar.selectbox("Fixed side", ["SS", "DS"], index=0)
    elif side_mode == "Embedded in another column":
        side_source = st.sidebar.selectbox("Column that contains side info", columns)
        side_source_letter = infer_letter_from_header(side_source) or col_index_to_letter(columns.index(side_source))
        st.sidebar.caption(f"Using column letter: {side_source_letter}")
        side_embedded_pattern = st.sidebar.text_input(
            "Optional: pattern in that column (ignored for now)", "", help="(Reserved for future use)"
        )

    # ---------- Group settings ----------
    st.sidebar.header("3. Grouping")

    group_mode = st.sidebar.radio("Group name mode", ["Fixed", "From another column"])
    group_fixed_name = None
    group_source_letter = None

    if group_mode == "Fixed":
        group_fixed_name = st.sidebar.text_input("Fixed group name", "UNGROUPED")
    else:
        group_source = st.sidebar.selectbox("Column that contains group / material info", columns)
        group_source_letter = infer_letter_from_header(group_source) or col_index_to_letter(columns.index(group_source))
        st.sidebar.caption(f"Using column letter: {group_source_letter}")

    # ---------- Material groups & DS materials ----------
    st.sidebar.header("4. Material groups & DS loading")

    mat_group_json = st.sidebar.text_area(
        "Material groups JSON",
        value=json.dumps(
            {
                "CORFLUTE_3MM": ["3mm corflute", "3 mm corflute", "3mm coreflute"],
                "SCREENBOARD_2MM": ["2mm screenboard", "2 mm screen board"],
                "POSTER_BOARD": ["400gsm artboard", "poster"],
                "WINDOW_SUPERCLING": ["jellyfish supercling"],
                "BANNER_SYNTHETIC": ["synthetic banner", "synthetic ds"],
                "FERROUS": ["ferrous substrate"],
                "ACM": ["acm"],
                "ACRYLIC": ["acrylic"],
                "BRAILLE": ["braille"],
                "STICKERS": ["sticker", "label"],
                "OTHER": ["misc"],
            },
            indent=2,
        ),
        height=250,
        help="Map group names to lists of text patterns that identify them.",
    )

    ds_materials_text = st.sidebar.text_area(
        "Double-sided (DS) material groups (one per line)",
        value="\\n".join(sorted(DEFAULT_DS_MATERIALS)),
        height=100,
    )

    ds_loading = st.sidebar.number_input(
        "DS loading (fraction)", min_value=0.0, max_value=1.0, value=DEFAULT_DS_LOADING, step=0.05
    )

    # ---------- Rates ----------
    st.sidebar.header("5. Rates per sqm")

    st.sidebar.markdown("Enter rates per sqm for each group. Leave blank if not used.")

    # Try to parse material groups JSON
    try:
        material_groups = json.loads(mat_group_json)
    except Exception as e:
        st.error(f"Invalid material groups JSON: {e}")
        return

    ds_materials = {line.strip() for line in ds_materials_text.splitlines() if line.strip()}
    st.sidebar.caption(f"DS loading applies to: {', '.join(sorted(ds_materials))}")

    all_group_names = sorted(set(material_groups.keys()) | set(DEFAULT_GROUP_NAMES) | ds_materials)

    material_rates = {}
    for g in all_group_names:
        rate = st.sidebar.text_input(f"Rate for {g} ($/sqm)", "")
        try:
            val = float(rate) if rate.strip() else np.nan
        except ValueError:
            val = np.nan
        material_rates[g] = val

    st.sidebar.header("6. Output mapping")

    sqm_col_letter = st.sidebar.text_input("SQM output column letter (row-wise)", "AC")
    price_col_letter = st.sidebar.text_input("Price output column letter (row-wise)", "AD")

    sqm_row_offset = st.sidebar.number_input(
        "SQM output row number (header row for sqm results)", min_value=1, value=171, step=1
    )
    price_row_offset = st.sidebar.number_input(
        "Price output row number (header row for price results)", min_value=1, value=172, step=1
    )

    start_data_row = st.sidebar.number_input(
        "First data row in Excel (0-based dataframe index -> this row)", min_value=1, value=5, step=1
    )

    if st.button("Run pricing"):
        if size_col == "(none)":
            st.error("You must choose a size column.")
            return
        if mat_col == "(none)":
            st.error("You must choose a material column.")
            return

        size_col_name = size_col if size_col != "(none)" else None
        mat_col_name = mat_col if mat_col != "(none)" else None
        qty_annum_col_name = qty_annum_col if qty_annum_col != "(none)" else None

        # Build pricing rows
        results = build_pricing_rows(
            df=df,
            size_col=size_col_name,
            mat_col=mat_col_name,
            qty_annum_col=qty_annum_col_name,
            side_mode=side_mode,
            side_fixed_value=side_fixed_value,
            side_embedded_pattern=side_embedded_pattern,
            side_source_letter=side_source_letter,
            qty_annum_default=qty_annum_default_val,
            default_side=default_side,
            group_mode=group_mode,
            group_fixed_name=group_fixed_name,
            group_source_letter=group_source_letter,
            ds_materials=ds_materials,
            ds_loading=ds_loading,
            material_groups=material_groups,
            material_rates=material_rates,
            qty_per_period_col=qty_per_period_col,
        )

        st.subheader("Calculated results (preview)")
        res_df = pd.DataFrame([r[1] for r in results])
        st.dataframe(res_df)

        # Write back to workbook
        try:
            out_bytes = write_results_to_workbook(
                wb_bytes=file_bytes,
                df_original=df,
                results=results,
                sqm_col_letter=sqm_col_letter,
                price_col_letter=price_col_letter,
                sqm_row_offset=sqm_row_offset,
                price_row_offset=price_row_offset,
                start_data_row=start_data_row,
            )
        except Exception as e:
            st.error(f"Failed to write results to workbook: {e}")
            return

        st.success("Pricing completed. Download updated workbook below.")
        st.download_button(
            label="Download updated Excel",
            data=out_bytes,
            file_name="tender_pricing_output.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


if __name__ == "__main__":
    main()
